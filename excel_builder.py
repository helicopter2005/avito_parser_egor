import re
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QMessageBox

def build_excel(self, data_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Оценка"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    align = Alignment(vertical="bottom", wrap_text=True)
    bold = Font(bold=True)

    headers = [
        "№",
        "Адрес",
        "Цена, руб",
        "Площадь, кв.м",
        "Цена за м², руб",
        "Этаж",
        "Площадь участка, кв.м",
        "Материал стен",
        "Год постройки",
        "Ссылка",
        "Описание",
        "Статус аналога"
    ]

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.border = border
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- ОСНОВНАЯ ТАБЛИЦА ---
    for idx, row in enumerate(data_rows, start=1):
        data = row["data"]
        is_analog = row["is_analog"]

        status_text = (
            "Выбран в качестве аналога"
            if is_analog else
            "Исключен по критерию"
        )

        try:
            ws.append([
                idx,
                data.get("address"),
                data.get("price"),
                data.get("area_m2"),
                data.get("price_per_m2"),
                int(data.get("params", {}).get("Этаж")) if str(data.get("params", {}).get("Этаж", "")).isdigit() else None,
                data.get("params", {}).get("Площадь участка"),
                data.get("params", {}).get("Материал стен"),
                data.get("params", {}).get("Год постройки"),
                data.get("url"),
                data.get("description"),
                status_text
            ])

            row_idx = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = align
                if col == 12 and is_analog:
                    cell.font = bold
        except Exception as e:
            QMessageBox.critical(self, "Ошибка при составлении основной таблицы", str(e))


    # --- ШИРИНА СТОЛБЦОВ ---
    widths = [10, 20, 10, 10, 10, 10, 10, 10, 14, 30, 65, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- ВТОРАЯ ТАБЛИЦА ---
    analogs = [r["data"] for r in data_rows if r["is_analog"]]

    if analogs:
        analogs_ws = wb.create_sheet(title="Аналоги")
        start_row = 1
        labels = [
            "Адрес",
            "Цена, руб",
            "Площадь, кв.м",
            "Цена за м², руб",
            "Этаж",
            "Площадь участка, кв.м",
            "Ссылка"
        ]

        for i, label in enumerate(labels):
            cell = analogs_ws.cell(row=start_row + i, column=1, value=label)
            cell.font = bold
            cell.border = border
            cell.alignment = align

        for col_idx, analog in enumerate(analogs, start=2):

            if analogs_ws.column_dimensions[get_column_letter(col_idx)].width < 20:
                analogs_ws.column_dimensions[get_column_letter(col_idx)].width = 20
            try:
                values = [
                    analog.get("address"),
                    analog.get("price"),
                    analog.get("area_m2"),
                    analog.get("price_per_m2"),
                    int(analog.get("params", {}).get("Этаж")) if str(analog.get("params", {}).get("Этаж", "")).isdigit() else None,
                    analog.get("params", {}).get("Площадь участка"),
                    analog.get("url"),
                ]

                for row_offset, value in enumerate(values):
                    cell = analogs_ws.cell(
                        row=start_row + row_offset,
                        column=col_idx,
                        value=value
                    )
                    cell.border = border
                    cell.alignment = align
            except Exception as e:
                QMessageBox.critical(self, "Ошибка при составлении таблицы с аналогами", str(e))

    return wb
